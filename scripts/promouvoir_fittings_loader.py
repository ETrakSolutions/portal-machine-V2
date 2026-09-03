# -*- coding: utf-8 -*-
"""Rend visibles en soumission les fittings des 101 chargeuses muettes.

CONTEXTE (Jacquot, 2026-09-03). Les fittings importes de « liste fitting sur
loader.xls » ne sortaient jamais d'une soumission de balance : la page de
soumission n'imprime que les items de kit en statut « Obligatoire », et il faut
en plus l'etiquette opt:"balance" pour qu'un item sorte avec la balance SEULE
(sans cette etiquette, la regle historique exige que le limiteur soit ON).
Les 15 machines-annees saisies a la main (Caterpillar 980 XE, John Deere
844 P-Tier, LiuGong 835HV) ont les deux et sortent correctement : elles servent
de modele ici.

REGLES ARBITREES PAR JACQUOT
  - lignes RACCORD -> statut r (obligatoire) + opt:"balance" ;
  - PN absent -> ligne obligatoire « Raccord a confirmer a la commande » ;
  - PLUSIEURS PN dans un seul champ -> UNE LIGNE PAR NUMERO (« une ligne par
    numero »), pour que chaque code soit collable dans la grille Epicor ;
  - lignes BOULONS -> NON TOUCHEES : elles n'ont aucun PN Epicor, et en
    obligatoire elles enverraient une ligne que personne ne peut commander.

ETIQUETAGE DES LIGNES DECOUPEES — le point delicat. La description du fichier
source agrege parfois PLUSIEURS lignes d'Excel avec des « ; » (le script
d'import faisait ' ; '.join). Apparier positionnellement les morceaux separes
par « / » aux numeros donnerait alors des etiquettes fausses : sur
« C02A-0021 / GPFS2406-1212-4 », la description « ORF #12 court parker ; Oring
Face # 12 court Parke ; ORF#12 long / ORF#12 long » se couperait n'importe
comment. On n'apparie donc QUE si la parenthese ne contient aucun « ; » ET se
coupe en exactement autant de morceaux qu'il y a de numeros. Sinon la ligne
porte son numero comme etiquette, et la description d'origine est reportee dans
l'Excel de controle pour arbitrage — jamais devinee.

QUANTITE : le fichier source n'en portait aucune. Les lignes restent donc a 1
(qty absent). Le 980 XE saisi a la main est a 2 : colonne « qty a confirmer »
dans l'Excel.

Ecriture COMPACTE (machines.json tient sur une ligne). --essai n'ecrit rien.
"""
import json
import io
import sys
import os
import re
import argparse

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PATH = os.path.join(REPO, 'data', 'machines.json')

# Code affiche pour une ligne sans numero de piece. Il paraît tel quel dans la
# soumission qui part chez le concessionnaire : donc lisible, et surtout PAS un
# faux numero de piece. Il doit rester identique a CODES_NON_COMMANDABLES dans
# js/kit-rules.js, qui l'ecarte du bloc Epicor.
CODE_A_CONFIRMER = 'À CONFIRMER'

ap = argparse.ArgumentParser()
ap.add_argument('--essai', action='store_true', help='n ecrit rien, affiche seulement')
ap.add_argument('--excel', default=os.path.join(REPO, 'scripts', 'Fittings_Loader_controle.xlsx'))
args = ap.parse_args()


def parenthese(desc):
    """Contenu de la DERNIERE parenthese ouvrante jusqu a la derniere fermante."""
    if not desc:
        return ''
    i, j = desc.find('('), desc.rfind(')')
    return desc[i + 1:j].strip() if (i >= 0 and j > i) else ''


def morceaux(txt):
    return [p.strip() for p in txt.split('/') if p.strip()]


def numeros(pn):
    """Numeros distincts, ordre d apparition conserve."""
    vus, out = set(), []
    for p in morceaux(pn):
        if p not in vus:
            vus.add(p)
            out.append(p)
    return out


def lignes_pour(c):
    """Les lignes RACCORD a produire pour une ligne source. Retourne
    (liste de lignes, etiquette du traitement applique)."""
    pn = (c.get('pn') or '').strip()
    desc = (c.get('desc') or '').strip() or 'Raccord hydraulique'
    nums = numeros(pn)
    base = {k: v for k, v in c.items() if k not in ('pn', 'desc', 'code', 'status', 'opt')}

    def ligne(code, d):
        r = dict(base)
        r.update({'code': code, 'pn': code, 'desc': d, 'status': 'r', 'opt': 'balance'})
        return r

    if not nums:
        det = parenthese(desc)
        d = 'Raccord hydraulique — numéro à confirmer à la commande'
        if det:
            d += ' (' + det + ')'
        r = dict(base)
        r.update({'code': CODE_A_CONFIRMER, 'pn': '', 'desc': d,
                  'desc_en': ('Hydraulic fitting — part number to be confirmed at order'
                              + ((' (' + det + ')') if det else '')),
                  'status': 'r', 'opt': 'balance'})
        return [r], 'sans PN -> a confirmer a la commande'

    if len(nums) == 1:
        return [ligne(nums[0], desc)], 'PN unique -> obligatoire + opt balance'

    det = parenthese(desc)
    parts = morceaux(det)
    if det and ';' not in det and len(parts) == len(nums):
        return ([ligne(n, 'Raccord hydraulique (' + p + ')') for n, p in zip(nums, parts)],
                'decoupe en %d, etiquettes appariees' % len(nums))
    return ([ligne(n, 'Raccord hydraulique (' + n + ')') for n in nums],
            'decoupe en %d, etiquettes = le numero (description source ambigue)' % len(nums))


d = json.load(open(PATH, encoding='utf-8'))
L = d['Loader']

rapport = []   # une ligne d Excel par machine-annee traitee
n_mach = n_avant = n_apres = 0
for b in sorted(L):
    if b.startswith('_'):
        continue
    for y in sorted(L[b]):
        for m in sorted(L[b][y]):
            sp = L[b][y][m]
            bom = sp.get('_bom')
            if not (isinstance(bom, dict) and isinstance(bom.get('_custom'), list)):
                continue
            cust = bom['_custom']
            if not any(c.get('code') == 'RACCORD' for c in cust):
                continue
            n_mach += 1
            neuf = []
            for c in cust:
                if c.get('code') != 'RACCORD':
                    neuf.append(c)          # BOULONS et tout le reste : intact
                    continue
                n_avant += 1
                prod, quoi = lignes_pour(c)
                n_apres += len(prod)
                neuf.extend(prod)
                rapport.append({
                    'Marque': b, 'Annee': y, 'Modele': m,
                    'PN source': c.get('pn') or '',
                    'Description source': c.get('desc') or '',
                    'Statut avant': c.get('status'),
                    'Traitement': quoi,
                    'Lignes produites': len(prod),
                    'Codes produits': ' | '.join(x['pn'] or '(aucun)' for x in prod),
                    'Etiquettes produites': ' | '.join(x['desc'] for x in prod),
                    'Qte a confirmer': 1,
                })
            bom['_custom'] = neuf

print('Machines-annees touchees          : %d' % n_mach)
print('Lignes RACCORD avant / apres      : %d -> %d' % (n_avant, n_apres))
from collections import Counter
for quoi, n in Counter(r['Traitement'] for r in rapport).most_common():
    print('   %-64s %d' % (quoi, n))

if args.essai:
    print('\n--essai : rien n a ete ecrit.')
else:
    tmp = PATH + '.tmp'
    with open(tmp, 'w', encoding='utf-8', newline='') as f:
        json.dump(d, f, ensure_ascii=False, separators=(',', ':'))
    json.load(open(tmp, encoding='utf-8'))          # relu avant de remplacer
    os.replace(tmp, PATH)
    print('\ndata/machines.json reecrit (compact, une ligne).')

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = 'Fittings Loader'
    cols = ['Marque', 'Annee', 'Modele', 'PN source', 'Description source',
            'Statut avant', 'Traitement', 'Lignes produites', 'Codes produits',
            'Etiquettes produites', 'Qte a confirmer']
    ws.append(cols)
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='1F4E79')
        c.alignment = Alignment(vertical='center')
    ambre = PatternFill('solid', fgColor='FFF2CC')
    for r in rapport:
        ws.append([r[c] for c in cols])
        if 'ambigue' in r['Traitement'] or 'a confirmer' in r['Traitement']:
            for c in ws[ws.max_row]:
                c.fill = ambre
    for col, w in zip('ABCDEFGHIJK', [18, 8, 16, 34, 52, 12, 58, 10, 40, 52, 14]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    wb.save(args.excel)
    print('Excel de controle : %s (%d lignes, les ambigues en ambre)'
          % (args.excel, len(rapport)))
except ImportError:
    print('openpyxl absent : pas d Excel de controle.')
