"""
Exporte la BD complete (machines + kits + notes + flags + admin) vers un fichier
Excel multi-onglets pret a importer dans Google Sheets.

Replique la logique computeDefaultBom de database.html en Python pour pre-remplir
les statuts R / J / NA par modele.
"""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

ROOT = Path(__file__).parent.parent
SRC_MACHINES = ROOT / "data" / "machines.json"
SRC_INSTALLED = ROOT / "data" / "installed_machines.json"
DST = ROOT / "data" / "Portal_DB_Test_v11.xlsx"

TYPE_TO_TAB = {
    "Excavatrice": "Excavatrice",
    "Pompe a Beton": "PompeBeton",
    "Grue Mobile": "GrueMobile",
    "Camion Girafe (Boom Truck)": "BoomTruck",
    "Telehandler": "Telehandler",
    "Foreuse": "Foreuse",
    "Camion Vacuum": "CamionVacuum",
    "Retrocaveuse": "Retrocaveuse",
}

# Replique de DRAIN_PREFIXES dans database.html / app.js
DRAIN_PREFIXES = [
    "CX80", "CX145", "CX170", "CX210", "CX220", "CX245", "CX300", "CX350", "CX380",
    "CX490", "145 D", "308", "315", "316", "320", "336", "440", "450", "M318",
    "DX190", "DX235", "BX190",
    "ZX210LC", "EX200", "ZX130-6", "ZX190", "ZX350", "ZX490", "ZX50U", "ZX75US", "ZX245",
    "245X",
    "135", "200CLC", "210G", "210P", "210 P", "245 P", "245P", "330X", "350", "410",
    "470G", "490D", "130P",
    "SK210",
    "PC78", "PC138", "PC200", "PC290",
    "R 920 K", "R920", "R 936", "R936",
    "145 X4", "145X4", "160 X4", "160X4", "170X4", "190", "245X4", "300 X4", "300X4",
    "350 X4", "350X4", "355 X4", "355X4", "490 X4", "490X4",
    "TB210", "TW65",
    "EC160", "EC330", "EC360", "EC550", "235",
    "EZ36",
]

EXC_BOM_CODES = ["0000", "0001", "0002", "0004", "0005", "0008", "0009", "0070", "0304"]
EXC_BOM_NAMES = {
    "0000": "Cabine", "0001": "Hauteur", "0002": "Rotation", "0004": "Mini exc",
    "0005": "Multi Axes", "0008": "Swing boom", "0009": "Drain hyd",
    "0070": "Boite GC", "0304": "Cremaillere",
}

# Note: Cameras, IDC et Creusage sont des options de soumission (pas BOM par machine)
# -> geres dans l'onglet _Listes (section Options_Soumission)
EXC_OPT_CODES = []

# Colonnes reserve pour pieces a venir
RESERVE_COLS = [
    ("RES-1", "Reserve 1"),
    ("RES-2", "Reserve 2"),
    ("RES-3", "Reserve 3"),
    ("RES-4", "Reserve 4"),
    ("RES-5", "Reserve 5"),
]

POMPE_BOM_CODES = ["0200", "0203", "0201", "0202", "0204", "0205", "0206", "0207", "0208", "0209"]
POMPE_BOM_NAMES = {
    "0200": "Avec coffre", "0203": "Sans coffre",
    "0201": "Hauteur", "0202": "Rotation",
    "0204": "4 sections", "0205": "5 sections", "0206": "6 sections",
    "0207": "Rot. cylindre", "0208": "Inclinometre", "0209": "Reel 15M",
}

HARNAIS_BY_FAB = {
    "HITACHI_7": ("Z03B-0121", "Hitachi -7"),
    "HITACHI_56": ("Z03B-0031", "Hitachi -5/-6"),
    "JOHN DEERE": ("Z03B-0031", "Hitachi/JD"),
    "KOMATSU": ("Z03B-0032", "Komatsu"),
    "DOOSAN": ("Z03B-0033", "Doosan"),
    "DEVELON": ("Z03B-0033", "Doosan"),
    "VOLVO": ("Z03B-0034", "Volvo"),
    "LINK-BELT": ("Z03B-0041", "Link-Belt/Case"),
    "CASE": ("Z03B-0041", "Link-Belt/Case"),
    "CATERPILLAR": ("Z03B-0080", "Caterpillar"),
    "CAT": ("Z03B-0080", "Caterpillar"),
}
HARNAIS_DEFAULT = ("Z03B-0043", "Generique")

# Couleurs
HEADER_FILL = PatternFill("solid", fgColor="1F2A37")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
ZEBRA_FILL = PatternFill("solid", fgColor="F3F4F6")
KEY_FILL = PatternFill("solid", fgColor="DBEAFE")
RED_FILL = PatternFill("solid", fgColor="FECACA")
YELLOW_FILL = PatternFill("solid", fgColor="FEF3C7")
NA_FILL = PatternFill("solid", fgColor="F3F4F6")
THIN = Side(border_style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load_machines():
    with open(SRC_MACHINES, "r", encoding="utf-8") as f:
        return json.load(f)


def load_installed():
    try:
        with open(SRC_INSTALLED, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return []


def parse_kg(value):
    if not value:
        return 0
    import re
    m = re.match(r"(\d[\d\s]*)", str(value))
    if not m:
        return 0
    try:
        return int(m.group(1).replace(" ", ""))
    except Exception:
        return 0


def get_harnais(fab, modele):
    fab_up = fab.upper()
    mod = modele or ""
    if fab_up == "HITACHI":
        is7 = "-7" in mod
        is56 = "-5" in mod or "-6" in mod
        if is7 and not is56:
            return HARNAIS_BY_FAB["HITACHI_7"]
        return HARNAIS_BY_FAB["HITACHI_56"]
    for key, value in HARNAIS_BY_FAB.items():
        if key in ("HITACHI_7", "HITACHI_56"):
            continue
        if key in fab_up:
            return value
    return HARNAIS_DEFAULT


def compute_default_bom_excavatrice(specs, fab, modele):
    poids_kg = parse_kg(specs.get("Poids operationnel (kg / lbs)", ""))
    has_swing = (specs.get("Swing boom", "") or "").lower() == "oui"
    is_mini = 0 < poids_kg < 5000
    mod_up = (modele or "").upper()
    is_drain = any(mod_up.startswith(p.upper()) for p in DRAIN_PREFIXES)
    return {
        "0000": "R",
        "0001": "J",
        "0002": "J",
        "0004": "R" if is_mini else "NA",
        "0005": "J",
        "0008": "J" if has_swing else "NA",
        "0009": "R" if is_drain else "NA",
        "0070": "NA",
        "0304": "R" if mod_up == "TB216" else "NA",
    }


def compute_default_bom_pompe(specs):
    nb_str = specs.get("Nombre de sections", "")
    try:
        sec = int(str(nb_str).strip().split()[0]) if nb_str else 0
    except Exception:
        sec = 0
    return {
        "0200": "NA", "0203": "NA",
        "0201": "J", "0202": "J",
        "0204": "R" if sec >= 4 else "NA",
        "0205": "R" if sec >= 5 else "NA",
        "0206": "R" if sec >= 6 else "NA",
        "0207": "NA", "0208": "NA", "0209": "NA",
    }


def collect_specs_rows(data, type_name):
    """Collecte les lignes pour un type de machine.
    Pour Excavatrice et Pompe a Beton: ajoute les colonnes BOM/Kits inline.
    Pour les autres: juste specs + Note Technicien.
    """
    fabs = data.get(type_name, {})
    rows = []
    seen_fields = []
    for fab, annees in fabs.items():
        if not isinstance(annees, dict):
            continue
        for annee, modeles in annees.items():
            if not isinstance(modeles, dict):
                continue
            for modele, specs in modeles.items():
                if not isinstance(specs, dict):
                    continue
                row = {"Fabricant": fab, "Annee": annee, "Modele": modele}
                for k, v in specs.items():
                    row[k] = v
                    if k not in seen_fields:
                        seen_fields.append(k)
                # Calcul BOM defauts selon le type
                if type_name == "Excavatrice":
                    bom = compute_default_bom_excavatrice(specs, fab, modele)
                    hcode, hname = get_harnais(fab, modele)
                    for code in EXC_BOM_CODES:
                        row[f"{code} {EXC_BOM_NAMES[code]}"] = bom.get(code, "NA")
                    for code, name in EXC_OPT_CODES:
                        row[f"{code} {name}"] = "NA"
                    for code, name in RESERVE_COLS:
                        row[f"{code} {name}"] = ""
                    # Concatene code + nom pour que la description apparaisse dans le dropdown
                    row["Harnais"] = f"{hcode} - {hname}"
                    row["Source BOM"] = "Defaut"
                elif type_name == "Pompe a Beton":
                    bom = compute_default_bom_pompe(specs)
                    for code in POMPE_BOM_CODES:
                        row[f"{code} {POMPE_BOM_NAMES[code]}"] = bom.get(code, "NA")
                    for code, name in RESERVE_COLS:
                        row[f"{code} {name}"] = ""
                    row["Source BOM"] = "Defaut"
                # Notes inline
                row["Note Technicien"] = ""
                row["Note Tech Auteur"] = ""
                row["Note Tech Date"] = ""
                row["Actif"] = "Oui"
                rows.append(row)

    # Ordre des colonnes selon le type
    if type_name == "Excavatrice":
        kit_cols = [f"{c} {EXC_BOM_NAMES[c]}" for c in EXC_BOM_CODES]
        kit_cols += [f"{c} {n}" for c, n in EXC_OPT_CODES]
        kit_cols += [f"{c} {n}" for c, n in RESERVE_COLS]
        kit_cols += ["Harnais", "Source BOM"]
    elif type_name == "Pompe a Beton":
        kit_cols = [f"{c} {POMPE_BOM_NAMES[c]}" for c in POMPE_BOM_CODES]
        kit_cols += [f"{c} {n}" for c, n in RESERVE_COLS]
        kit_cols += ["Source BOM"]
    else:
        kit_cols = []

    columns = ["Fabricant", "Annee", "Modele"] + seen_fields + kit_cols + [
        "Note Technicien", "Note Tech Auteur", "Note Tech Date", "Actif"
    ]
    rows.sort(key=lambda r: (r["Fabricant"], str(r.get("Modele", "")), str(r.get("Annee", ""))))
    return columns, rows


def is_kit_column(col_name):
    """Detecte si une colonne est une colonne de kit (BOM) — commence par un code numerique."""
    if not col_name:
        return False
    parts = col_name.split(" ", 1)
    head = parts[0]
    # Code BOM (4 chiffres) ou code optionnel (1300-XXXX, 1000-XXXX) ou reserve (RES-X)
    return head.isdigit() or head.startswith("1300-") or head.startswith("1000-") or head.startswith("RES-")


def write_specs_sheet(wb, tab_name, columns, rows, n_key_cols=3):
    ws = wb.create_sheet(tab_name)
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BORDER

    # Indices des colonnes kit (pour DataValidation et formatage conditionnel)
    kit_col_indices = [
        idx for idx, name in enumerate(columns, start=1) if is_kit_column(name)
    ]

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            val = row.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = BORDER

            if col_idx <= n_key_cols:
                cell.fill = KEY_FILL
            elif is_kit_column(col_name):
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(bold=True)
            elif col_name == "Actif":
                cell.alignment = Alignment(horizontal="center")
                if val == "Non":
                    cell.fill = PatternFill("solid", fgColor="FECACA")
                    cell.font = Font(color="991B1B", bold=True)
            elif row_idx % 2 == 0:
                cell.fill = ZEBRA_FILL

    # Ajout des regles de mise en forme conditionnelle sur les colonnes kit
    # (la couleur suit la valeur quand l'utilisateur edite)
    if kit_col_indices and len(rows) > 0:
        first_row = 2
        last_row = len(rows) + 1
        red_fill = PatternFill("solid", fgColor="FECACA")
        red_font = Font(bold=True, color="991B1B")
        yellow_fill = PatternFill("solid", fgColor="FEF3C7")
        yellow_font = Font(bold=True, color="92400E")
        gray_fill = PatternFill("solid", fgColor="F3F4F6")
        gray_font = Font(color="9CA3AF")

        for col_idx in kit_col_indices:
            col_letter = get_column_letter(col_idx)
            range_ref = f"{col_letter}{first_row}:{col_letter}{last_row}"
            # R = rouge
            ws.conditional_formatting.add(
                range_ref,
                CellIsRule(operator="equal", formula=['"R"'], fill=red_fill, font=red_font)
            )
            # J = jaune
            ws.conditional_formatting.add(
                range_ref,
                CellIsRule(operator="equal", formula=['"J"'], fill=yellow_fill, font=yellow_font)
            )
            # NA = gris
            ws.conditional_formatting.add(
                range_ref,
                CellIsRule(operator="equal", formula=['"NA"'], fill=gray_fill, font=gray_font)
            )

    # Liste deroulante (DataValidation) sur les colonnes kit
    if kit_col_indices and len(rows) > 0:
        dv = DataValidation(
            type="list",
            formula1='"R,J,NA"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Valeur invalide",
            error="Choisis R (obligatoire), J (optionnel), NA (non applicable) ou laisse vide.",
            showInputMessage=True,
            promptTitle="Statut du kit",
            prompt="R = obligatoire (rouge)\nJ = optionnel (jaune)\nNA = non applicable\nVide = reserve non utilisee",
        )
        for col_idx in kit_col_indices:
            col_letter = get_column_letter(col_idx)
            dv.add(f"{col_letter}{2}:{col_letter}{len(rows) + 1}")
        ws.add_data_validation(dv)

    # Liste deroulante pour Harnais (Code + Description concatenes)
    if len(rows) > 0:
        for ci, col_name in enumerate(columns, start=1):
            if col_name == "Harnais":
                dv_h = DataValidation(
                    type="list",
                    formula1="=Harnais_Codes!$G$2:$G$100",  # colonne G = Affichage
                    allow_blank=True,
                    showErrorMessage=False,
                    showInputMessage=True,
                    promptTitle="Harnais",
                    prompt="Format: <Code> - <Description>\nGere dans l'onglet Harnais_Codes.",
                )
                col_letter = get_column_letter(ci)
                dv_h.add(f"{col_letter}{2}:{col_letter}{len(rows) + 1}")
                ws.add_data_validation(dv_h)
            elif col_name == "Source BOM":
                dv_src = DataValidation(
                    type="list",
                    formula1='"Defaut,Manuel,Override"',
                    allow_blank=True,
                )
                col_letter = get_column_letter(ci)
                dv_src.add(f"{col_letter}{2}:{col_letter}{len(rows) + 1}")
                ws.add_data_validation(dv_src)
            elif col_name == "Actif":
                dv_a = DataValidation(
                    type="list",
                    formula1='"Oui,Non"',
                    allow_blank=False,
                    showInputMessage=True,
                    promptTitle="Modele actif?",
                    prompt="Oui = visible dans le portail\nNon = masque (mais conserve)",
                )
                col_letter = get_column_letter(ci)
                dv_a.add(f"{col_letter}{2}:{col_letter}{len(rows) + 1}")
                ws.add_data_validation(dv_a)

    for col_idx, col_name in enumerate(columns, start=1):
        if is_kit_column(col_name):
            ws.column_dimensions[get_column_letter(col_idx)].width = 14
        else:
            max_len = max(
                [len(str(col_name))] + [len(str(r.get(col_name, ""))) for r in rows[:200]]
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_len + 2), 40)
    ws.freeze_panes = ws.cell(row=2, column=n_key_cols + 1)
    last_col = get_column_letter(len(columns))
    ws.auto_filter.ref = f"A1:{last_col}{max(len(rows) + 1, 1)}"
    ws.row_dimensions[1].height = 28


def write_kit_excavatrice(wb, data):
    """Onglet Kits_Excavatrice : 1 ligne / modele + colonnes BOM avec defauts calcules."""
    ws = wb.create_sheet("Kits_Excavatrice")
    headers = ["Fabricant", "Annee", "Modele"]
    headers += [f"{c} {EXC_BOM_NAMES[c]}" for c in EXC_BOM_CODES]
    headers += [f"{c} {n}" for c, n in EXC_OPT_CODES]
    headers += [f"{c} {n}" for c, n in RESERVE_COLS]
    headers += ["Harnais Code", "Harnais Nom", "Source"]

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BORDER

    fabs = data.get("Excavatrice", {})
    rows = []
    for fab, annees in fabs.items():
        if not isinstance(annees, dict):
            continue
        for annee, modeles in annees.items():
            if not isinstance(modeles, dict):
                continue
            for modele, specs in modeles.items():
                if not isinstance(specs, dict):
                    continue
                bom = compute_default_bom_excavatrice(specs, fab, modele)
                hcode, hname = get_harnais(fab, modele)
                rows.append({
                    "Fabricant": fab, "Annee": annee, "Modele": modele,
                    "bom": bom, "hcode": hcode, "hname": hname,
                })
    rows.sort(key=lambda r: (r["Fabricant"], str(r["Modele"]), str(r["Annee"])))

    for row_idx, r in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=r["Fabricant"]).fill = KEY_FILL
        ws.cell(row=row_idx, column=2, value=r["Annee"]).fill = KEY_FILL
        ws.cell(row=row_idx, column=3, value=r["Modele"]).fill = KEY_FILL
        for ci, code in enumerate(EXC_BOM_CODES, start=4):
            v = r["bom"].get(code, "NA")
            cell = ws.cell(row=row_idx, column=ci, value=v)
            cell.alignment = Alignment(horizontal="center")
            if v == "R":
                cell.fill = RED_FILL
                cell.font = Font(bold=True, color="991B1B")
            elif v == "J":
                cell.fill = YELLOW_FILL
                cell.font = Font(bold=True, color="92400E")
            else:
                cell.fill = NA_FILL
                cell.font = Font(color="9CA3AF")
        # Colonnes optionnelles (cameras, IDC) — vides par defaut
        opt_start = 4 + len(EXC_BOM_CODES)
        for ci, _ in enumerate(EXC_OPT_CODES, start=opt_start):
            cell = ws.cell(row=row_idx, column=ci, value="NA")
            cell.alignment = Alignment(horizontal="center")
            cell.fill = NA_FILL
            cell.font = Font(color="9CA3AF")
        # Colonnes reserve — vides
        res_start = opt_start + len(EXC_OPT_CODES)
        for ci, _ in enumerate(RESERVE_COLS, start=res_start):
            cell = ws.cell(row=row_idx, column=ci, value="")
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill("solid", fgColor="FAFAFA")
        # Harnais et source
        h_start = res_start + len(RESERVE_COLS)
        ws.cell(row=row_idx, column=h_start, value=r["hcode"])
        ws.cell(row=row_idx, column=h_start + 1, value=r["hname"])
        ws.cell(row=row_idx, column=h_start + 2, value="Defaut")
        for c in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=c).border = BORDER

    for col_idx, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, len(h) + 2), 24)
    ws.freeze_panes = ws.cell(row=2, column=4)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    ws.row_dimensions[1].height = 28


def write_kit_pompe(wb, data):
    ws = wb.create_sheet("Kits_PompeBeton")
    headers = ["Fabricant", "Annee", "Modele"]
    headers += [f"{c} {POMPE_BOM_NAMES[c]}" for c in POMPE_BOM_CODES]
    headers += [f"{c} {n}" for c, n in RESERVE_COLS]
    headers += ["Source"]

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BORDER

    fabs = data.get("Pompe a Beton", {})
    rows = []
    for fab, annees in fabs.items():
        if not isinstance(annees, dict):
            continue
        for annee, modeles in annees.items():
            if not isinstance(modeles, dict):
                continue
            for modele, specs in modeles.items():
                if not isinstance(specs, dict):
                    continue
                bom = compute_default_bom_pompe(specs)
                rows.append({
                    "Fabricant": fab, "Annee": annee, "Modele": modele, "bom": bom
                })
    rows.sort(key=lambda r: (r["Fabricant"], str(r["Modele"]), str(r["Annee"])))

    for row_idx, r in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=r["Fabricant"]).fill = KEY_FILL
        ws.cell(row=row_idx, column=2, value=r["Annee"]).fill = KEY_FILL
        ws.cell(row=row_idx, column=3, value=r["Modele"]).fill = KEY_FILL
        for ci, code in enumerate(POMPE_BOM_CODES, start=4):
            v = r["bom"].get(code, "NA")
            cell = ws.cell(row=row_idx, column=ci, value=v)
            cell.alignment = Alignment(horizontal="center")
            if v == "R":
                cell.fill = RED_FILL
                cell.font = Font(bold=True, color="991B1B")
            elif v == "J":
                cell.fill = YELLOW_FILL
                cell.font = Font(bold=True, color="92400E")
            else:
                cell.fill = NA_FILL
                cell.font = Font(color="9CA3AF")
        # Colonnes reserve
        res_start = 4 + len(POMPE_BOM_CODES)
        for ci, _ in enumerate(RESERVE_COLS, start=res_start):
            cell = ws.cell(row=row_idx, column=ci, value="")
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill("solid", fgColor="FAFAFA")
        ws.cell(row=row_idx, column=res_start + len(RESERVE_COLS), value="Defaut")
        for c in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=c).border = BORDER

    for col_idx, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, len(h) + 2), 24)
    ws.freeze_panes = ws.cell(row=2, column=4)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    ws.row_dimensions[1].height = 28


def write_simple_sheet(wb, name, headers, rows):
    ws = wb.create_sheet(name)
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, v in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=v)
            cell.border = BORDER
            if row_idx % 2 == 0:
                cell.fill = ZEBRA_FILL
    for col_idx, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, len(str(h)) + 2), 40)
    ws.freeze_panes = ws.cell(row=2, column=1)
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{max(len(rows) + 1, 1)}"
    ws.row_dimensions[1].height = 24
    return ws


def add_dropdown_from_listes(ws, header_name, list_col_letter, max_rows=200, prompt=""):
    """Ajoute un menu deroulant a la colonne avec ce header, qui reference _Listes."""
    # Trouver l'indice de colonne par header
    target_col = None
    for col_idx in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col_idx).value == header_name:
            target_col = col_idx
            break
    if not target_col:
        return
    target_letter = get_column_letter(target_col)
    dv = DataValidation(
        type="list",
        formula1=f"=_Listes!${list_col_letter}$2:${list_col_letter}$15",
        allow_blank=True,
        showInputMessage=bool(prompt),
        promptTitle=header_name,
        prompt=prompt,
    )
    dv.add(f"{target_letter}2:{target_letter}{max_rows}")
    ws.add_data_validation(dv)


def write_listes_sheet(wb):
    """Onglet _Listes : centralise les menus deroulants gerable par l'utilisateur.
    Chaque colonne represente une liste; les onglets concernes y font reference
    via DataValidation.
    """
    ws = wb.create_sheet("_Listes")

    # Structure: colonnes A, C, E, G, I, K = listes; B, D, F, H, J = descriptions
    lists = {
        "Severite": [
            ("info", "Bandeau bleu, info utile"),
            ("warning", "Bandeau jaune, verification necessaire"),
            ("danger", "Bandeau rouge, attention bloquante"),
        ],
        "Trigger_Type": [
            ("always", "Toujours afficher"),
            ("single", "Si CE kit est active"),
            ("all_of", "Si TOUS ces kits actives (ET)"),
            ("any_of", "Si AU MOINS UN kit active (OU)"),
            ("none_of", "Si AUCUN de ces kits"),
        ],
        "Roles": [
            ("super_admin", "Tous droits"),
            ("administrateur", "Admin general"),
            ("vente_interne", "Equipe vente"),
            ("technicien", "Installateurs / techs"),
            ("ingenierie", "Equipe ingenierie"),
            ("distributeur", "Distributeurs externes"),
            ("dealer", "Concessionnaires"),
        ],
        "Source_BOM": [
            ("Defaut", "BOM calcule a partir des specs"),
            ("Manuel", "Modifie manuellement"),
            ("Override", "Override admin"),
        ],
        "Statut_Demande": [
            ("Nouveau", "Demande recue, pas traitee"),
            ("En cours", "Specs en train d'etre completees"),
            ("Complete", "Toutes les specs remplies"),
            ("Rejete", "Demande rejetee"),
        ],
        "Oui_Non": [
            ("Oui", ""),
            ("Non", ""),
        ],
        "Options_Soumission": [
            ("1300-0001 - Camera Recul", "Camera de recul standard"),
            ("1300-0003 - Camera Quad", "Camera 4 vues"),
            ("1300-0004 - Camera 360", "Vue panoramique 360"),
            ("1000-0004 - IDC Lock Valve", "Indicateur de charge avec lock valve"),
            ("1500-0006 - Limiteur Creusage", "Limiteur de profondeur de creusage"),
        ],
    }

    # Header row
    title_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    desc_fill = PatternFill("solid", fgColor="F3F4F6")
    col_idx = 1
    for list_name, items in lists.items():
        # Header de la liste (fusion sur 2 colonnes)
        c1 = ws.cell(row=1, column=col_idx, value=list_name)
        c1.font = title_font
        c1.fill = HEADER_FILL
        c1.alignment = Alignment(horizontal="center", vertical="center")
        c1.border = BORDER
        c2 = ws.cell(row=1, column=col_idx + 1, value="Description")
        c2.font = title_font
        c2.fill = HEADER_FILL
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c2.border = BORDER
        # Items (avec lignes vides en bas pour ajout futur)
        for ri, (val, desc) in enumerate(items, start=2):
            cell_val = ws.cell(row=ri, column=col_idx, value=val)
            cell_val.font = Font(bold=True)
            cell_val.border = BORDER
            cell_desc = ws.cell(row=ri, column=col_idx + 1, value=desc)
            cell_desc.fill = desc_fill
            cell_desc.border = BORDER
        # 5 lignes vides pour extension
        for ri in range(len(items) + 2, len(items) + 7):
            ws.cell(row=ri, column=col_idx, value="").border = BORDER
            ws.cell(row=ri, column=col_idx + 1, value="").border = BORDER

        ws.column_dimensions[get_column_letter(col_idx)].width = 18
        ws.column_dimensions[get_column_letter(col_idx + 1)].width = 32
        col_idx += 3  # gap d'1 colonne entre les listes

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # Note explicative
    note_row = 20
    ws.cell(row=note_row, column=1, value="COMMENT UTILISER CET ONGLET").font = Font(bold=True, size=12)
    notes = [
        "1. Pour ajouter une nouvelle valeur dans un menu deroulant, ajoute-la dans la colonne correspondante (lignes vides en bas).",
        "2. Pour retirer une valeur: efface la cellule (ne laisse pas de trou — remonte les valeurs).",
        "3. Les onglets qui utilisent ces listes detecteront automatiquement les changements.",
        "4. Pour les harnais (plus complexes), aller dans l'onglet Harnais_Codes.",
        "5. Pour les codes BOM (1500-XXXX, etc.), aller dans l'onglet BOM_Codes.",
    ]
    for i, n in enumerate(notes, start=1):
        ws.cell(row=note_row + i, column=1, value=n)


def write_readme(wb, stats):
    ws = wb.create_sheet("_Lisez-moi", 0)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 60

    rows = [
        ["Portal Machine — Base de donnees TEST", ""],
        ["", ""],
        ["Genere le", "2026-04-29"],
        ["Source", "machines.json + installed_machines.json + logique BOM database.html"],
        ["", ""],
        ["[ STRUCTURE ]", ""],
        ["", ""],
        ["DONNEES MACHINES (1 onglet par type)", ""],
        ["Excavatrice / PompeBeton / GrueMobile / BoomTruck", "Specs + Note Technicien inline (auteur, date)"],
        ["Telehandler / Foreuse / CamionVacuum / Retrocaveuse", "Specs + Note Technicien inline (auteur, date)"],
        ["", ""],
        ["KITS / BOM (fusionnes dans les onglets machines)", ""],
        ["Excavatrice", "Inclut colonnes BOM: 0000-0304, cameras, IDC, RES-1..5, Harnais"],
        ["PompeBeton", "Inclut colonnes BOM: 0200-0209 + RES-1..5"],
        ["Couleurs BOM", "Rouge=R obligatoire | Jaune=J optionnel | Gris=NA | Vide=reserve"],
        ["", ""],
        ["ASSOCIATIF", ""],
        ["Notes", "Notes texte par modele — toutes roles avec writeNotes (admin/ing/technicien)"],
        ["Warnings_Soumission", "Avertissement affiche sur la page soumission pour ce modele"],
        ["Flags", "Drapeaux BOM (signalements de defaut)"],
        ["Demandes_creation", "Log des nouvelles demandes 'Autre modele'"],
        ["Changelog", "Audit des modifications"],
        ["", ""],
        ["REFERENCE", ""],
        ["_Listes", "GESTION CENTRALISEE des menus deroulants (Severite, Trigger, Roles...)"],
        ["BOM_Codes", "Liste des codes BOM avec descriptions"],
        ["Harnais_Codes", "Codes harnais — ajouter une ligne pour creer un nouveau harnais"],
        ["Emails", "Listes courriels admin (kit, sales, target, notes)"],
        ["Vendeurs", "Liste vendeurs"],
        ["Users", "Utilisateurs autorises (lecture seule recommandee)"],
        ["Roles", "Permissions par role"],
        ["Machines_Installees", "Historique des machines installees"],
        ["", ""],
        ["[ COULEURS DANS Kits_* ]", ""],
        ["Rouge (R)", "Kit obligatoire — defaut a partir des specs"],
        ["Jaune (J)", "Kit optionnel"],
        ["Gris (NA)", "Non applicable"],
        ["", ""],
        ["[ TRIGGERS Warnings_Soumission ]", ""],
        ["always", "Toujours afficher quand ce modele est selectionne"],
        ["single", "Si CE kit est active (Trigger Codes = 1 code)"],
        ["all_of", "Si TOUS ces kits sont actives (codes separes par virgule)"],
        ["any_of", "Si AU MOINS UN de ces kits est active"],
        ["none_of", "Si AUCUN de ces kits est active"],
        ["", ""],
        ["Severite", "info (bleu) | warning (jaune) | danger (rouge)"],
        ["Joker '(tous)'", "Dans Fabricant/Modele/Annee, signifie 'tous les modeles de ce type'"],
        ["", ""],
        ["[ INVENTAIRE ]", ""],
    ]
    for r in rows:
        ws.append(r)
    for typ, n in stats.items():
        ws.append([typ, f"{n} fiches"])
    ws.append(["", ""])
    ws.append(["[ UTILISATION ]", ""])
    ws.append(["1.", "Importer ce .xlsx dans Google Drive"])
    ws.append(["2.", "Clic droit -> Ouvrir avec Google Sheets"])
    ws.append(["3.", "Fichier -> Enregistrer en tant que Google Sheets"])
    ws.append(["4.", "Tester filtres, vues, modifications"])

    for cell in ws["A"]:
        cell.font = Font(bold=True)


def main():
    print("Lecture des sources...")
    machines = load_machines()
    installed = load_installed()

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Onglets specs (un par type)
    stats = {}
    for type_name, tab_name in TYPE_TO_TAB.items():
        if type_name not in machines:
            continue
        cols, rows = collect_specs_rows(machines, type_name)
        write_specs_sheet(wb, tab_name, cols, rows)
        stats[type_name] = len(rows)
        print(f"  {tab_name}: {len(rows)} lignes")

    # Note: les Kits sont maintenant fusionnes dans les onglets Excavatrice et PompeBeton
    # (colonnes inline apres les specs, avant les notes techniciens)

    # Onglets associatifs (vides, prets a recevoir)
    write_simple_sheet(wb, "Notes",
        ["Type", "Fabricant", "Annee", "Modele", "Note", "Auteur (role)", "Date"],
        [])
    warnings_ws = write_simple_sheet(wb, "Warnings_Soumission",
        [
            "Type", "Fabricant", "Annee", "Modele",
            "Trigger Type", "Trigger Codes",
            "Warning (texte affiche)", "Severite",
            "Auteur", "Date"
        ],
        [
            # Exemples illustrant chaque type de trigger
            ["Excavatrice", "(tous)", "(tous)", "(tous)",
             "single", "1000-0004",
             "IDC: vérifier que le modèle a une lock valve compatible",
             "warning", "(exemple systeme)", "2026-04-29"],

            ["Excavatrice", "Caterpillar", "(tous)", "320",
             "all_of", "0009, 1000-0004",
             "Drain hydraulique + IDC: prévoir adaptateur Z03B-CAT spécifique",
             "danger", "(exemple)", "2026-04-29"],

            ["Excavatrice", "Hitachi", "(tous)", "ZX135US-7",
             "always", "",
             "Harnais Hitachi-7 obligatoire (Z03B-0121)",
             "info", "(exemple)", "2026-04-29"],

            ["Excavatrice", "(tous)", "(tous)", "(tous)",
             "any_of", "1300-0001, 1300-0003, 1300-0004",
             "Caméra installée: vérifier alimentation 24V disponible",
             "info", "(exemple)", "2026-04-29"],

            ["Pompe a Beton", "(tous)", "(tous)", "(tous)",
             "single", "0206",
             "6 sections: harnais long requis (Z03B-PMP-LONG)",
             "warning", "(exemple)", "2026-04-29"],

            ["Excavatrice", "Takeuchi", "(tous)", "TB216",
             "none_of", "0009",
             "Crémaillère prévue mais sans drain: confirmer specs avant install",
             "warning", "(exemple)", "2026-04-29"],
        ])
    # Menus deroulants references vers _Listes
    add_dropdown_from_listes(warnings_ws, "Trigger Type", "C",
                             prompt="always | single | all_of | any_of | none_of")
    add_dropdown_from_listes(warnings_ws, "Severite", "A",
                             prompt="info | warning | danger")

    # Prompt informatif sur Trigger Codes (pas de validation stricte car format CSV)
    for col_idx in range(1, warnings_ws.max_column + 1):
        if warnings_ws.cell(row=1, column=col_idx).value == "Trigger Codes":
            col_letter = get_column_letter(col_idx)
            dv_codes = DataValidation(
                type="custom",
                formula1='TRUE',
                showInputMessage=True,
                promptTitle="Trigger Codes",
                prompt=("Code(s) BOM separes par virgule.\n"
                        "Ex: 1000-0004 (single)\n"
                        "Ex: 0009, 1000-0004 (all_of)\n"
                        "Voir onglet BOM_Codes pour la liste complete."),
            )
            dv_codes.add(f"{col_letter}2:{col_letter}200")
            warnings_ws.add_data_validation(dv_codes)
            break
    write_simple_sheet(wb, "Flags",
        ["Type", "Fabricant", "Modele", "Annee", "FlaggedBy", "FlaggedAt", "Note", "Active"],
        [])
    demandes_ws = write_simple_sheet(wb, "Demandes_creation",
        ["Date", "User", "Type", "Fabricant", "Annee", "Modele", "Status"],
        [])
    add_dropdown_from_listes(demandes_ws, "Status", "I",
                             prompt="Nouveau | En cours | Complete | Rejete")
    write_simple_sheet(wb, "Changelog",
        ["Timestamp", "Utilisateur", "Type", "Fabricant", "Modele", "Annee", "Valeur precedente", "Nature changement"],
        [])

    # Reference
    write_simple_sheet(wb, "BOM_Codes",
        ["Code", "Nom", "Categorie", "Description", "Numero piece"],
        [
            ["1500-0000", "Cabine", "Excavatrice", "Sortie cabine principale", "1500-0000"],
            ["1500-0001", "Hauteur", "Excavatrice", "Limiteur de hauteur", "1500-0001"],
            ["1500-0002", "Rotation", "Excavatrice", "Limiteur de rotation", "1500-0002"],
            ["1500-0004", "Mini exc", "Excavatrice", "Mini excavatrice (poids < 5000 kg)", "1500-0004"],
            ["1500-0005", "Multi Axes", "Excavatrice", "Limiteur multi-axes", "1500-0005"],
            ["1500-0008", "Swing boom", "Excavatrice", "Swing boom", "1500-0008"],
            ["1500-0009", "Drain hyd", "Excavatrice", "Drain hydraulique (modeles specifiques)", "1500-0009"],
            ["1500-0070", "Boite GC", "Excavatrice", "Boite de commande GC", "1500-0070"],
            ["1500-0304", "Cremaillere", "Excavatrice", "Cremaillere (TB216 specifique)", "1500-0304"],
            ["1500-0200", "Avec coffre", "Pompe Beton", "Console avec coffre", "1500-0200"],
            ["1500-0201", "Hauteur", "Pompe Beton", "Limiteur hauteur pompe", "1500-0201"],
            ["1500-0202", "Rotation", "Pompe Beton", "Limiteur rotation pompe", "1500-0202"],
            ["1500-0203", "Sans coffre", "Pompe Beton", "Console sans coffre", "1500-0203"],
            ["1500-0204", "4 sections", "Pompe Beton", "4 sections (>=4)", "1500-0204"],
            ["1500-0205", "5 sections", "Pompe Beton", "5 sections (>=5)", "1500-0205"],
            ["1500-0206", "6 sections", "Pompe Beton", "6 sections (>=6)", "1500-0206"],
            ["1500-0207", "Rot. cylindre", "Pompe Beton", "Rotation cylindre", "1500-0207"],
            ["1500-0208", "Inclinometre", "Pompe Beton", "Inclinometre", "1500-0208"],
            ["1500-0209", "Reel 15M", "Pompe Beton", "Reel 15M", "1500-0209"],
            ["1300-0001", "Camera Recul", "Options Soumission", "Camera de recul", "1300-0001"],
            ["1300-0003", "Camera Quad", "Options Soumission", "Camera quad-vue", "1300-0003"],
            ["1300-0004", "Camera 360", "Options Soumission", "Camera 360 degres", "1300-0004"],
            ["1000-0004", "IDC Lock Valve", "Options Soumission", "Indicateur de charge", "1000-0004"],
            ["1500-0006", "Limiteur Creusage", "Options Soumission", "Limiteur profondeur creusage", "1500-0006"],
        ])

    # Onglet _Listes : gestion centralisee des menus deroulants
    write_listes_sheet(wb)

    harnais_ws = write_simple_sheet(wb, "Harnais_Codes",
        ["Code", "Nom", "Fabricants applicables", "Numero piece", "Actif", "Notes", "Affichage (auto)"],
        [
            ["Z03B-0031", "Hitachi -5/-6 / John Deere", "Hitachi (-5/-6), John Deere", "Z03B-0031", "Oui", "", ""],
            ["Z03B-0032", "Komatsu", "Komatsu", "Z03B-0032", "Oui", "", ""],
            ["Z03B-0033", "Doosan / Develon", "Doosan, Develon", "Z03B-0033", "Oui", "", ""],
            ["Z03B-0034", "Volvo", "Volvo CE", "Z03B-0034", "Oui", "", ""],
            ["Z03B-0041", "Link-Belt / Case", "Link-Belt, Case", "Z03B-0041", "Oui", "", ""],
            ["Z03B-0080", "Caterpillar", "Caterpillar / CAT", "Z03B-0080", "Oui", "", ""],
            ["Z03B-0100", "Caterpillar (ECU)", "Caterpillar avec ECU", "Z03B-0100", "Oui", "", ""],
            ["Z03B-0121", "Hitachi -7", "Hitachi (-7)", "Z03B-0121", "Oui", "", ""],
            ["Z03B-0043", "Generique", "Autres / inconnus", "Z03B-0043", "Oui", "", ""],
            # Lignes vides pour ajout futur
            ["", "", "", "", "", "", ""],
            ["", "", "", "", "", "", ""],
            ["", "", "", "", "", "", ""],
            ["", "", "", "", "", "", ""],
            ["", "", "", "", "", "", ""],
        ])
    # Colonne G "Affichage (auto)" : formule qui concatene Code + " - " + Nom
    # (utilisee comme source du dropdown Harnais dans Excavatrice)
    for r in range(2, 15):
        cell = harnais_ws.cell(row=r, column=7,
                               value=f'=IF(A{r}="","",A{r}&" - "&B{r})')
        cell.fill = PatternFill("solid", fgColor="F0F9FF")
        cell.font = Font(italic=True, color="0369A1")
        cell.border = BORDER
    harnais_ws.column_dimensions["G"].width = 45

    # Emails (placeholders — a remplir depuis l'API)
    write_simple_sheet(wb, "Emails",
        ["Liste", "Adresse", "Notes"],
        [
            ["kit_emails", "(a remplir)", "Recoit les demandes de kit machine"],
            ["sales_emails", "(a remplir)", "Equipe vente"],
            ["target_emails", "(a remplir)", "Cibles soumissions"],
            ["notes_emails", "(a remplir)", "Notifications nouvelles notes"],
        ])

    write_simple_sheet(wb, "Vendeurs",
        ["Nom", "Email", "Region", "Actif"],
        [])

    users_ws = write_simple_sheet(wb, "Users",
        ["Username", "Nom complet", "Email", "Role", "Actif"],
        [])
    add_dropdown_from_listes(users_ws, "Role", "E",
                             prompt="Role utilisateur (gere dans _Listes ou Roles)")
    add_dropdown_from_listes(users_ws, "Actif", "K", prompt="Oui / Non")

    write_simple_sheet(wb, "Roles",
        ["Role", "createAccount", "modifBom", "kitMachineAccess", "soumissionAccess",
         "shareAccess", "writeNotes", "modifAccounts", "machineAccess", "databaseAccess", "flagBom"],
        [
            ["super_admin", True, True, True, True, True, True, True, True, True, True],
            ["administrateur", True, True, True, True, True, True, True, True, True, True],
            ["vente_interne", True, False, True, True, True, False, False, True, False, False],
            ["technicien", False, False, False, False, False, True, False, True, False, False],
            ["distributeur", False, False, True, True, False, False, False, True, False, False],
            ["dealer", False, False, True, True, False, False, False, True, False, False],
            ["ingenierie", False, True, False, False, False, True, False, True, True, True],
        ])

    # Machines installees (depuis installed_machines.json)
    inst_rows = []
    for entry in installed:
        inst_rows.append([
            entry.get("marque", ""),
            entry.get("modele", ""),
            entry.get("annee", ""),
            entry.get("date_installation", ""),
            entry.get("client", ""),
            entry.get("vendeur", ""),
            entry.get("notes", ""),
        ])
    write_simple_sheet(wb, "Machines_Installees",
        ["Marque", "Modele", "Annee", "Date installation", "Client", "Vendeur", "Notes"],
        inst_rows)
    print(f"  Machines_Installees: {len(inst_rows)} lignes")

    # Readme en premier
    write_readme(wb, stats)

    DST.parent.mkdir(parents=True, exist_ok=True)
    wb.save(DST)
    print(f"\nFichier genere: {DST}")
    print(f"Total onglets: {len(wb.sheetnames)}")
    print(f"Onglets: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
