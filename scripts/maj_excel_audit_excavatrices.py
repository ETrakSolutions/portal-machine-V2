# -*- coding: utf-8 -*-
"""Marque dans l'Excel d'audit ce qui a ete importe en BD le jour dit.

Ajoute deux colonnes a la feuille « Manquants a valider » :
  - « Importe le »  : date d'import (vide si non importe)
  - « Note import » : nom nettoye, classe corrigee, annees rognees, raison
                      de non-import.
"""
import json, os, re, sys, collections
from datetime import date

import openpyxl

sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from import_excavatrices_manquantes import (FAB_MAP, clean_modele, poids_kg_list,
                                            db_weight_points, classe_voisins, LOT_PATH)

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(os.path.dirname(ROOT), 'Audit_Excavatrices_Manquantes_2015-2026.xlsx')
DATE = sys.argv[1] if len(sys.argv) > 1 else date.today().isoformat()

db = json.load(open(os.path.join(ROOT, 'data', 'machines.json'), encoding='utf-8'))
ex = db['Excavatrice']
pts = db_weight_points(ex)

present = collections.defaultdict(set)      # (fab, modele) -> annees en BD
for fab, yy in ex.items():
    if fab.startswith('_'):
        continue
    for y, mm in yy.items():
        for m in mm:
            present[(fab, m)].add(y)

# Decision du 2026-08-05 : les 4 Bobcat laisses vides sont finalement retenus.
DECIDES_O = {('Bobcat', 'E25'), ('Bobcat', 'E45'), ('Bobcat', 'E48'), ('Bobcat', 'E80')}

wb = openpyxl.load_workbook(XLSX)
ws = wb['Manquants a valider']
for _r in range(2, ws.max_row + 1):
    if (ws.cell(row=_r, column=1).value, ws.cell(row=_r, column=2).value) in DECIDES_O:
        if not str(ws.cell(row=_r, column=9).value or '').strip():
            ws.cell(row=_r, column=9, value='o')
head = [c.value for c in ws[1]]
for col in ('Importe le', 'Note import'):
    if col not in head:
        ws.cell(row=1, column=len(head) + 1, value=col)
        head.append(col)
ci, cn = head.index('Importe le') + 1, head.index('Note import') + 1

n_imp = n_non = 0
for r in range(2, ws.max_row + 1):
    fab_src = ws.cell(row=r, column=1).value
    mod_src = ws.cell(row=r, column=2).value
    if not fab_src:
        continue
    fab = FAB_MAP.get(fab_src, fab_src)
    mod = clean_modele(mod_src)
    annees_src = str(ws.cell(row=r, column=3).value or '')
    valide = str(ws.cell(row=r, column=9).value or '').strip().lower() in ('o', 'oui')

    notes = []
    if mod != mod_src:
        notes.append('nom nettoye -> "%s"' % mod)

    en_bd = present.get((fab, mod))
    if not valide:
        ws.cell(row=r, column=cn, value='non valide (colonne « A ajouter ? » vide)')
        continue
    if not en_bd:
        raison = 'non importe'
        a = re.findall(r'\d{4}', annees_src)
        if a:
            want = set(str(y) for y in range(int(a[0]), int(a[-1]) + 1))
            dispo = set(ex.get(fab, {}).keys())
            if not (want & dispo):
                raison = ('non importe : le fabricant ne couvre que %s-%s en BD'
                          % (min(dispo), max(dispo)))
        notes.append(raison)
        n_non += 1
    else:
        n_imp += 1
        ws.cell(row=r, column=ci, value=DATE)
        notes.append('en BD : %s-%s' % (min(en_bd), max(en_bd)))
        # classe reellement inscrite
        y0 = sorted(en_bd)[0]
        cl = ex[fab][y0][mod].get('Classe machine')
        if cl != str(ws.cell(row=r, column=4).value):
            notes.append('classe corrigee : %s -> %s' % (ws.cell(row=r, column=4).value, cl))
        a = re.findall(r'\d{4}', annees_src)
        if a:
            want = set(str(y) for y in range(int(a[0]), int(a[-1]) + 1))
            miss = sorted(want - en_bd)
            if miss:
                notes.append('annees non creees (absentes du fabricant) : %s' % ','.join(miss))
    ws.cell(row=r, column=cn, value=' ; '.join(notes))

wb.save(XLSX)
print('Excel mis a jour : %s' % XLSX)
print('  lignes marquees importees : %d' % n_imp)
print('  lignes validees non importees : %d' % n_non)
