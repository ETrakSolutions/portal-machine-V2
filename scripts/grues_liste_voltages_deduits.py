# -*- coding: utf-8 -*-
"""Liste les voltages encore issus d'une DEDUCTION (jamais confirmes en fiche).

Recalcule l'etat REEL de la base : les 17 deductions Link-Belt ayant ete
retirees, il reste les autres. Produit un tableau lisible + un Excel de suivi.
"""
import json, os, sys, collections

sys.stdout.reconfigure(encoding='utf-8')
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CHAMP = 'Voltage machine (V/type)'
VIDE = ('', 'a completer', 'à compléter', 'n/d', 'nd', '-')

liste = json.load(open(os.path.join(ROOT, 'scripts', 'data',
                                    'grues_voltage_a_verifier.json'), encoding='utf-8'))
gm = json.load(open(os.path.join(ROOT, 'data', 'machines.json'), encoding='utf-8'))['Grue Mobile']

restants = []
for x in liste:
    f, m = x['marque'], x['modele']
    annees = sorted(y for y in gm.get(f, {}) if m in gm[f][y])
    if not annees:
        continue
    actuel = str(gm[f][annees[0]][m].get(CHAMP) or '').strip()
    if actuel.lower() in VIDE:
        continue          # deduction retiree (Link-Belt)
    restants.append({**x, 'voltage_en_base': actuel, 'nb_entrees': len(annees)})

print('VOLTAGES ENCORE DEDUITS, NON VERIFIES EN FICHE : %d modeles '
      '(%d entrees annee-modele)\n'
      % (len(restants), sum(r['nb_entrees'] for r in restants)))

par_marque = collections.defaultdict(list)
for r in restants:
    par_marque[r['marque']].append(r)

for f in sorted(par_marque, key=lambda k: -len(par_marque[k])):
    rs = par_marque[f]
    prio = rs[0]['priorite']
    print('=== %s — %d modeles%s ==='
          % (f, len(rs), '   [PRIORITE HAUTE : non-uniformite prouvee]' if prio == 'haute' else ''))
    for r in sorted(rs, key=lambda x: x['modele']):
        print('   %-24s %-26s %-10s %-9s %s'
              % (r['modele'][:24], r['famille'][:26], r['annees'],
                 r['capacite'], r['voltage_en_base']))
    print()

# Excel de suivi
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Voltages a verifier'
    entetes = ['Marque', 'Modele', 'Famille', 'Annees', 'Capacite max',
               'Voltage deduit (en base)', 'Priorite', 'Voltage confirme',
               'Source', 'Verifie le']
    ws.append(entetes)
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='2F5496')
    jaune = PatternFill('solid', fgColor='FFF2CC')
    for r in sorted(restants, key=lambda x: (x['priorite'] != 'haute', x['marque'], x['modele'])):
        ws.append([r['marque'], r['modele'], r['famille'], r['annees'], r['capacite'],
                   r['voltage_en_base'], r['priorite'], '', '', ''])
        if r['priorite'] == 'haute':
            for c in ws[ws.max_row]:
                c.fill = jaune
    for col, w in zip('ABCDEFGHIJ', (20, 26, 28, 12, 14, 24, 11, 18, 40, 12)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    p = os.path.join(os.path.dirname(ROOT), 'Grues_Voltages_A_Verifier.xlsx')
    wb.save(p)
    print('Excel de suivi : %s' % p)
except ImportError:
    print('(openpyxl absent : pas d Excel genere)')
