# -*- coding: utf-8 -*-
"""Cloture des 4 dernieres lignes « En attente » de l'audit Grue Mobile.

Les 4 restantes sont des Demag TOUT-TERRAIN, toutes notees « Rebadge Tadano ».
Elles tombent sous la decision d'audit deja prise par l'utilisateur : « Demag
non distinct -> les chenilles CC sont repliees sous Tadano, les tout-terrain
Demag sont ecartees (Tadano rebadgees) ». On enregistre donc l'exclusion, avec
sa raison, dans la colonne Decision.

Aucune de ces 4 machines n'est en BD -> aucun changement cote portail.
"""
import os, sys
from datetime import date
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(os.path.dirname(ROOT), 'Audit_Grues_Mobiles_2026-07.xlsx')
DATE = sys.argv[1] if len(sys.argv) > 1 else date.today().isoformat()

wb = openpyxl.load_workbook(XLSX)
ws = wb['Audit grues mobiles']
head = [str(c.value) if c.value is not None else '' for c in ws[1]]
i_dec = head.index('Decision (O/N)') + 1
i_note = head.index('Note / doublon') + 1
i_marque, i_modele = head.index('Marque') + 1, head.index('Modele') + 1

n = 0
for r in range(2, ws.max_row + 1):
    if 'attente' in str(ws.cell(row=r, column=i_dec).value or ''):
        ws.cell(row=r, column=i_dec,
                value='Exclu - Demag AT = Tadano rebadgee (regle audit, %s)' % DATE)
        note = str(ws.cell(row=r, column=i_note).value or '')
        ws.cell(row=r, column=i_note,
                value=(note + ' | Clos le %s : jamais en BD, aucun changement portail.' % DATE).strip(' |'))
        print('  exclu : %s %s' % (ws.cell(row=r, column=i_marque).value,
                                   ws.cell(row=r, column=i_modele).value))
        n += 1
wb.save(XLSX)
print('%d ligne(s) closes. Plus aucune ligne « En attente ».' % n)
