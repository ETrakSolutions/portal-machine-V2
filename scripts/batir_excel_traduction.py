# -*- coding: utf-8 -*-
"""Assemble l'Excel de validation des traductions anglaises des textes libres de la BD.

CONTEXTE. Le mecanisme bilingue est livre depuis le 2026-09-02 (_notes_en /
_warning_en / desc_en, avec repli FR<->EN) mais le remplissage etait entier. Le
volume annonce — « ~3 290 textes » — comptait les OCCURRENCES : en textes
DISTINCTS il n'y en a que 244, pour 3 580 occurrences. Un seul texte (« Coupure
electrique... ») en occupe 1 116 a lui seul.

Cet Excel est la piece a valider AVANT toute ecriture dans la BD. Une ligne par
texte distinct : le francais, l'anglais propose, le nombre de machines touchees,
et une colonne « Anglais corrige » que Jacquot remplit s'il n'est pas d'accord.
Rien n'est ecrit dans machines.json ni dans les overrides par ce script.

CONTROLES AUTOMATIQUES (colonne « Signalement »), parce qu'un Excel de 244 lignes
ne se relit pas a l'oeil :
  - traduction manquante ou identique au francais ;
  - numeros de piece presents en francais mais absents de l'anglais (une
    reference perdue en traduction est une commande fausse) ;
  - noms propres de produits/personnes disparus (PRAN, Guide Pro, Keven...) ;
  - le marqueur de source [Source: ...] present d'un cote seulement ;
  - notes internes non redigees, a reecrire ou effacer plutot qu'a traduire.
"""
import json
import os
import io
import re
import sys
import importlib.util

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
# Les lots de traduction et l extraction FR vivent DANS le depot : le script doit
# pouvoir etre rejoue par n importe qui, sur n importe quel poste.
SC = os.path.join(REPO, 'scripts', 'traductions')
SORTIE = os.path.join(REPO, 'scripts', 'Traduction_textes_BD.xlsx')


def charger(nom):
    chemin = os.path.join(SC, nom)
    spec = importlib.util.spec_from_file_location(nom[:-3], chemin)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


mods = [charger('trad_lot%d.py' % i) for i in (1, 2, 3, 4)]
TRAD, A_REVOIR = {}, set()
for m in mods:
    doubles = set(TRAD) & set(m.TRAD)
    assert not doubles, 'id en double entre deux lots : %s' % sorted(doubles)
    TRAD.update(m.TRAD)
    A_REVOIR |= getattr(m, 'A_REVOIR', set())

textes = json.load(open(os.path.join(SC, 'textes_fr.json'), encoding='utf-8'))
manquants = [o['id'] for o in textes if o['id'] not in TRAD]
assert not manquants, 'textes sans traduction : %s' % manquants
orphelins = set(TRAD) - {o['id'] for o in textes}
assert not orphelins, 'traductions sans texte source : %s' % sorted(orphelins)

# Un « numero de piece » : au moins une lettre ET un chiffre, ou un code e-Trak.
RX_PN = re.compile(r'\b(?:[A-Z]{1,4}\d{2,}[A-Z0-9-]*|[A-Z]\d{2}[A-Z]-\d{4}|'
                   r'\d{4}-\d{4}|GPFS\d+[\w-]*|ASX?-\d+-?\w*|TN\d+\w*)\b')
NOMS = ['PRAN', 'Guide Pro', 'Limit Pro', 'Limitronic', 'Danfoss', 'Axiomatic', 'Omron',
        'Tyco', 'Keven', 'Mathieu', 'Pierre', 'Éric', 'ISOGARDE', 'Deutsch', 'Parker',
        'Hydro', 'HIAB', 'ROTOBEC', 'Gageport', 'Kobelco', 'SANY', 'Link-Belt', 'Link Belt']


def signaler(fr, en, tid):
    s = []
    if not en.strip():
        s.append('traduction VIDE')
        return s
    if en.strip() == fr.strip() and len(fr) > 30:
        s.append('anglais identique au francais')
    perdus = sorted(set(RX_PN.findall(fr)) - set(RX_PN.findall(en)))
    if perdus:
        s.append('numero(s) perdu(s) : ' + ', '.join(perdus))
    nperdus = [n for n in NOMS if n.lower() in fr.lower() and n.lower() not in en.lower()]
    if nperdus:
        s.append('nom(s) propre(s) perdu(s) : ' + ', '.join(nperdus))
    src_fr = 'ProgressionLive' in fr
    src_en = 'ProgressionLive' in en
    if src_fr != src_en:
        s.append('marqueur de source present d un seul cote')
    if tid in A_REVOIR:
        s.append('NOTE INTERNE non redigee — a reecrire ou effacer, pas a traduire')
    return s


lignes = []
for o in textes:
    en = TRAD[o['id']]
    lignes.append({
        'id': o['id'],
        'Machines': o['occ'],
        'Champ': o['champs'],
        'Exemple de machine': o['exemple'],
        'Francais (BD)': o['fr'],
        'Anglais propose': en,
        'Anglais corrige': '',
        'Signalement': ' | '.join(signaler(o['fr'], en, o['id'])),
    })

n_sig = sum(1 for l in lignes if l['Signalement'])
print('244 attendus, %d assembles.' % len(lignes))
print('Occurrences couvertes : %d' % sum(l['Machines'] for l in lignes))
print('Lignes portant un signalement : %d' % n_sig)
for l in lignes:
    if l['Signalement']:
        print('   #%03d  %-56s  %s' % (l['id'], l['Francais (BD)'][:54], l['Signalement']))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

wb = Workbook()
ws = wb.active
ws.title = 'Textes a valider'
cols = ['id', 'Machines', 'Champ', 'Exemple de machine', 'Francais (BD)',
        'Anglais propose', 'Anglais corrige', 'Signalement']
ws.append(cols)
for c in ws[1]:
    c.font = Font(bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='1F4E79')
    c.alignment = Alignment(vertical='center')
ambre = PatternFill('solid', fgColor='FFF2CC')
asaisir = PatternFill('solid', fgColor='EAF3FB')
for l in lignes:
    ws.append([l[c] for c in cols])
    r = ws.max_row
    for c in ws[r]:
        c.alignment = Alignment(vertical='top', wrap_text=True)
    ws.cell(row=r, column=7).fill = asaisir       # colonne a remplir
    if l['Signalement']:
        for c in ws[r]:
            c.fill = ambre
for col, w in zip('ABCDEFGH', [6, 10, 14, 34, 68, 68, 40, 40]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = 'E2'
ws.auto_filter.ref = ws.dimensions

# Deuxieme onglet : le vocabulaire retenu, pour que la relecture soit rapide.
v = wb.create_sheet('Vocabulaire')
v.append(['Francais', 'Anglais retenu', 'Provenance'])
for c in v[1]:
    c.font = Font(bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='1F4E79')
VOC = [
    ('limiteur de portée', 'range limiter', 'js/translations.js (interface)'),
    ('limiteur', 'limiter', 'js/translations.js'),
    ('coupure / harnais de coupure', 'cut-off / cut-off harness', 'js/translations.js'),
    ('guide de creusage', 'digging guide', 'js/translations.js'),
    ('indicateur de charge (IDC)', 'load indicator (IDC)', 'js/translations.js'),
    ('hauteur / rotation', 'height / rotation', 'js/translations.js'),
    ('multi-axe', 'multi-axis', 'js/translations.js'),
    ('rotation crémaillère', 'rack rotation', 'js/translations.js'),
    ('balance', 'scale', 'js/translations.js'),
    ('inclinomètre', 'inclinometer', 'js/translations.js'),
    ('capteur de proximité / prox', 'proximity sensor', 'usage e-Trak'),
    ('godet', 'bucket', 'usage e-Trak'),
    ('couronne (de rotation)', 'ring gear / slew ring', 'usage e-Trak'),
    ('équerre / bracket de prox', 'bracket', 'usage e-Trak'),
    ('attache rapide / quick attach', 'quick coupler / quick attach', 'usage e-Trak'),
    ('réel', 'reel', 'usage e-Trak'),
    ('bras de sécurité / de coupure', 'safety arm / cut-off arm', 'usage e-Trak'),
    ('bons de travail ProgressionLive', 'ProgressionLive work orders', 'usage e-Trak'),
    ('noms de produits (PRAN, Guide Pro, Limit Pro)', 'inchangés', 'noms propres'),
    ('numéros de pièce', 'inchangés', 'references Epicor'),
]
for r in VOC:
    v.append(list(r))
for col, w in zip('ABC', [42, 40, 32]):
    v.column_dimensions[col].width = w
v.freeze_panes = 'A2'

wb.save(SORTIE)
print()
print('Excel : %s' % SORTIE)
print('  onglet « Textes a valider » : %d lignes, colonne G a remplir si desaccord' % len(lignes))
print('  onglet « Vocabulaire »      : %d termes ancres sur l interface' % len(VOC))
print('AUCUNE ecriture dans machines.json ni dans les overrides.')
