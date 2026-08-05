# -*- coding: utf-8 -*-
"""Etat reel du chantier Grue Mobile : decisions de l'Excel d'audit vs contenu
de la BD, et mesure des specs restant a completer."""
import json, os, sys, collections

import openpyxl

sys.stdout.reconfigure(encoding='utf-8')
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(os.path.dirname(ROOT), 'Audit_Grues_Mobiles_2026-07.xlsx')

wb = openpyxl.load_workbook(XLSX, read_only=True)
print('feuilles :', wb.sheetnames)
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))
head = [str(c) if c is not None else '' for c in rows[0]]
print('colonnes :', head)
idx = {h: i for i, h in enumerate(head)}

col_stat = next((h for h in head if 'tatut' in h or 'Verdict' in h or 'verdict' in h), None)
col_dec = next((h for h in head if 'cision' in h), None)
print('\ncolonne statut :', col_stat, '| colonne decision :', col_dec)
if col_stat:
    print('repartition statut :', collections.Counter(
        str(r[idx[col_stat]]) for r in rows[1:] if r[idx[col_stat]] is not None))
if col_dec:
    print('repartition decision :', collections.Counter(
        str(r[idx[col_dec]]) for r in rows[1:]))
    print('\nlignes SANS decision :')
    n = 0
    for r in rows[1:]:
        if not str(r[idx[col_dec]] or '').strip():
            n += 1
            print('   ', ' | '.join(str(r[i]) for i in range(min(4, len(r)))))
    print('   total sans decision :', n)

# --- BD
db = json.load(open(os.path.join(ROOT, 'data', 'machines.json'), encoding='utf-8'))
gm = db['Grue Mobile']
mods = {(f, m) for f in gm if not f.startswith('_') for y in gm[f] for m in gm[f][y]}
ent = sum(1 for f in gm if not f.startswith('_') for y in gm[f] for m in gm[f][y])
print('\nBD Grue Mobile : %d modeles distincts, %d entrees, %d marques'
      % (len(mods), ent, len([f for f in gm if not f.startswith('_')])))
print('Manitex 45110T en BD :', any(f == 'Manitex' and m == '45110T' for f, m in mods))

# --- specs a completer
champs = collections.Counter()
tot = 0
for f in gm:
    if f.startswith('_'):
        continue
    for y in gm[f]:
        for m, v in gm[f][y].items():
            if not isinstance(v, dict):
                continue
            tot += 1
            for k, val in v.items():
                if k.startswith('_') or k in ('Flag', 'Image'):
                    continue
                s = str(val or '').strip()
                if s == '' or s.lower() in ('a completer', 'à compléter', 'a compléter'):
                    champs[k] += 1
print('\nspecs vides / « A completer » (sur %d entrees) :' % tot)
for k, n in champs.most_common():
    print('   %-42s %5d  (%.0f %%)' % (k, n, 100.0 * n / tot))
